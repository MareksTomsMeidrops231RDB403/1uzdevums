#Mareks Toms Meidrops, Grupa 16, 231RDB403
from openpyxl import Workbook, load_workbook
wb = load_workbook('test1.xlsx')
ws = wb.active
max_row = ws.max_row
print(max_row)
skaits = 0
for i in range(2, max_row+1):
    rate = ws['B'+str(i)].value
    hours = ws['C'+str(i)].value
    if (type(rate)!=str and type(hours)!=str):
        salary = hours*rate
        ws['D'+str(i)].value = salary
        #izvadu cilvēkus, kuriem ir lielāka mēneša alga par 3000
        if salary>=3000:
            print(salary)
            #skaitītājs, kas saskaita cilvēku skaitu, kuru alga ir lielāka par 3000
            skaits+=1
print('\nCilvēku skaits, kuru alga pārsniedz 3000 ir', skaits)
wb.save('test1.xlsx')
wb.close()